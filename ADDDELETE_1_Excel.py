import psycopg2
import dash
from dash import dcc, html, Input, Output, dash_table
from dash.dependencies import Input, Output,State, MATCH, ALL
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import dash_bootstrap_components as dbc  # Import dash-bootstrap-components
import subprocess
import os
import signal  # Import the signal module
from dash import callback_context
from dash.exceptions import PreventUpdate
from psycopg2 import sql
import threading
import time
import json
import queue
from Allocation_check_Excel import output_queue
from multiprocessing import Process, Queue, get_context
import webbrowser
from threading import Timer
from openpyxl import load_workbook
import sys


allocation_process=None

def launch_dashboard():

    allocation_process=None
    # Global variable to track interruption status
    allocation_interrupted = False
    Dash_time=None
    # Database connection details
    db_name = 'ProductDetails'
    db_username = 'PUser12'
    db_password = 'PSQL@123'
    db_host = 'localhost'
    db_port = '5432'
    global flag1
    # Function to start allocation process in a separate thread
    #file_path="Product Details_v2.xlsx"
    # Function to start allocation process in a separate thread

    full_file_path=r'D:\Downloads\Zulf\Product Details_v2.xlsx'
    # Function to fetch data from the database
    def fetch_data(full_file_path, sheet_name):
        
        try:
            data = pd.read_excel(full_file_path, sheet_name=sheet_name)
            
            return data
        except FileNotFoundError:
            print(f"Error: The file at {full_file_path} was not found.")
            return None
        except pd.errors.EmptyDataError:
            print(f"Error: The file at {full_file_path} is empty.")
            return None
        except pd.errors.ParserError:
            print(f"Error: The file at {full_file_path} does not appear to be a valid Excel file.")
            return None
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None



    # Function to fetch data from the Excel file
    def fetch_data_table(full_file_path, sheet_name, usecols=None):
        try:
            data = pd.read_excel(full_file_path, sheet_name=sheet_name, usecols=usecols)
            
            return data
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None
        
    # Function to fetch data from the database
    def fetch_data1(full_file_path, sheet_name):
        try:
            
            df = pd.read_excel(full_file_path, sheet_name=sheet_name)
            

            # Determine if any component of the product is not editable
            df['editable'] = df.groupby('Product Name')['Status'].transform(lambda x: not any(status in ['Completed', 'InProgress'] for status in x))

            # Filter the dataframe to include only products where all components are editable
            editable_df = df[df['editable']].drop(columns=['editable'])
            print(editable_df)
            return editable_df
            #return df
        except Exception as e:
            print(f"Error fetching data: {e}")
            return pd.DataFrame()

    # Function to convert time string to timedelta
    def time_to_timedelta2(t):
        try:
            if isinstance(t, datetime):
                return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
            if pd.isna(t) or t == "":
                return timedelta(0)
            # Ensure t is a string and in the format "HH:MM:SS"
            if isinstance(t, str) and ':' in t:
                h, m, s = map(int, t.split(":"))
                return timedelta(hours=h, minutes=m, seconds=s)
            else:
                # Handle unexpected input format or missing ':'
                raise ValueError(f"Unexpected format or missing ':' in input: {t}")
        except Exception as e:
            print(f"Error in time_to_timedelta2: {e}")
            return timedelta(0)  # or raise further or return appropriate default

    # Function to calculate utilization in minutes
    def calculate_utilization(t):
        total_seconds = t.total_seconds()
        return total_seconds / 60


    # Function to fetch previous data from the database
    def fetch_previous_data_from_db(full_file_path, sheet_name):
        try:
            return pd.read_excel(full_file_path, sheet_name=sheet_name)
            
        except Exception as e:
            print(f"Error fetching data from database: {e}")
            return []
        
    # Function to get the last unique ID from the database
    def get_last_unique_id(table_name,full_file_path):
        # Initialize variables to store last unique IDs
        last_unique_id = 0
        # Determine sheet name based on table_name
        if table_name == 'prodet':
            sheet_name = 'prodet' 
        elif 'Addln':
            sheet_name = 'Addln' 
        wb = load_workbook(full_file_path)
        sheet = wb[sheet_name]
        last_row = sheet.max_row
        # Assuming UniqueID is in the first column (A)
        last_unique_id = sheet.cell(row=last_row, column=1).value
        wb.close()
        return last_unique_id if last_unique_id else 0


    # Convert necessary fields to string
    def convert_data_for_json(data):
        for record in data:
            for key, value in record.items():
                if isinstance(value, pd.Timestamp):
                    record[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, (pd.Timedelta, pd.TimedeltaIndex)):
                    record[key] = str(value)
                elif pd.isna(value):  # Convert NaN to None for JSON serialization
                    record[key] = None
        return data



    # Function to read data from Excel files
    def fetch_data_from_excel(full_file_path, sheet_name):
        try:
            data = pd.read_excel(full_file_path, sheet_name=sheet_name)
            return data
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None
        
    # Function to get the full file path
    def get_file_path_1(filename):
        current_dir = os.path.dirname(__file__)
        return os.path.join(current_dir, filename)

    # File paths and sheet names
    #xcel_file_path = 'Product Details_v2.xlsx'
    #file_path = 'Product Details_v2.xlsx'
    full_file_path=r'Product Details_v2.xlsx'
    #full_file_path = get_file_path_1(file_path)
    print(f"Full file path: {full_file_path}")

    # Fetching raw data from Excel sheets
    prodet_df = fetch_data_from_excel(full_file_path, 'prodet')
    time_details_df = fetch_data_from_excel(full_file_path, 'Product Times')

    # Creating Product Lookup Table similar to the DB query
    if prodet_df is not None:
        product_df = prodet_df.groupby("Product Name").agg(
            Total_Quantity_Required=("Quantity Required", "sum"),
            Total_Components=("Quantity Required", "count"),
            Average_Run_Time=("Run Time (min/1000)", lambda x: round(x.mean(), 2))
        ).reset_index()
    else:
        product_df = pd.DataFrame()

    # Creating Component Lookup Table (already fetched as prodet_df)
    component_df = prodet_df if prodet_df is not None else pd.DataFrame()

    # Creating Machine Utilization Lookup Table similar to the DB query
    if prodet_df is not None:
        machine_df = prodet_df.groupby("Machine Number").agg(
            Total_Runtime=("Run Time (min/1000)", "sum"),
            Total_Components=("Quantity Required", "count"),
            Average_Cycle_Time=("Cycle Time (seconds)", "mean")
        ).reset_index()
    else:
        machine_df = pd.DataFrame()

    # Creating Order Details Lookup Table (fetched as time_details_df)
    order_details_df = time_details_df

    # Function to read data from the Excel sheet and process it
    def fetch_similarity_data():
        df = pd.read_excel(full_file_path, sheet_name='Similarity')
        print(df)
        # Apply transformation: replace 1 with ✓ and leave other cells as is
        df = df.map(lambda x: '✓' if x == 0 else ('' if x == 1 else x))
        return df


    initial_data = fetch_data1(full_file_path, "prodet").to_dict('records')
    initial_data_json = json.dumps(initial_data, default=str)

    # Initialize the starting time to 09:00:00
    start_time = datetime.combine(datetime.today(), datetime.min.time()) + timedelta(hours=9)

    # Define button style
    button_style = {
        'margin': '10px',
        'padding': '15px 30px',
        'font-size': '16px',
        'font-weight': 'bold',
        'border-radius': '8px',
        'background-color': '#3498db',
        'color': 'white',
        'border': 'none',
        'cursor': 'pointer',
        'transition': 'background-color 0.3s ease',
    }
    
    app = dash.Dash(__name__, external_stylesheets=[dbc.themes.SLATE])
    app.config.suppress_callback_exceptions = True

    # Define your CSS styles (assuming you have the CSS file as described)
    app.css.append_css({'external_url': '/assets/styles.css'})
    # Define layout
    app.layout = dbc.Container(
       
        style={'textAlign': 'left', 'width': '100%', 'margin': 'auto'},
        children=[
            html.Div(style={'height': '50px'}),
            html.H1('Dashboard - Production Analysis', style={'textAlign': 'center', 'marginBottom': '30px'}),

            dcc.Interval(
                id='interval-component-script',
                interval=1000,  # Update script control every second
                n_intervals=0,
                disabled=True  # Start disabled
            ),

            dbc.Row([
                dbc.Col([
                    html.Button('Read from Spreadsheet', id='read-button', n_clicks=0, style=button_style),
                    html.Button('Start', id='initialise-button', n_clicks=0, style=button_style),
                    html.Button('Reschedule', id='start-button', n_clicks=0, style=button_style),
                    html.Button('Pause', id='stop-button', n_clicks=0, style=button_style),
                    html.Button('Reset', id='reset-button', n_clicks=0, style=button_style),
                    html.Div(id='start-message', style={'marginLeft': '20px', 'color': 'green'})
                ], width=15, style={'textAlign': 'Right', 'margin': 'auto'})
            ]),
            
            # Modals for feedback messages
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='read-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-read-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="read-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='initialise-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-initialise-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="initialise-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='start-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-start-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="start-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='stop-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-stop-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="stop-modal",
                is_open=False,
            ),
            dbc.Modal(
                [
                    dbc.ModalHeader("Operation Status"),
                    dbc.ModalBody(id='reset-modal-body'),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-reset-modal", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="reset-modal",
                is_open=False,
            ),
            html.Div(
        id='completion-message',
        className='message-box',  # Apply the general box style
        style={'textAlign': 'center', 'marginTop': '20px', 'fontSize': '20px'}
    ),
    dcc.Interval(
        id='interval-component-data',
        interval=5000,  # Update every 5 seconds
        n_intervals=0
    ),

            dcc.Tabs(id='tabs', value='tab-input', children=[
                dcc.Tab(label='Product List change', value='tab-manage-products', children=[
                    dbc.Row([
                        dbc.Col(
                            dcc.Dropdown(
                                id='manage-dropdown',
                                options=[
                                    {'label': 'Add Product', 'value': 'add'},
                                    {'label': 'Delete Product', 'value': 'delete'},
                                    {'label': 'Swap Product', 'value': 'swap'}
                                ],
                                value='add',
                                placeholder='Select action',
                                style={'width': '200px'}
                            ),
                            width=3,
                            style={'padding': '20px'}
                        ),
                        dbc.Col(
                            html.Div(id='manage-content'),
                            width=9
                        )
                    ])
                ]),
                
                dcc.Tab(label='Product Catalogue', value='tab-2', children=[
                    html.H2('Below are the product details', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    html.Div([
                        dash_table.DataTable(
                            id='data-table',
                            columns=[],
                            data=[],
                            filter_action='native',
                            sort_action="native",
                            page_size=10,
                            style_table={'height': '900px', 'overflowY': 'auto', 'marginBottom': '20px'},
                            style_cell={
                                'textAlign': 'center',
                                'padding': '5px',
                                'backgroundColor': '#f9f9f9',
                                'border': '1px solid black',
                                'minWidth': '120px', 'maxWidth': '150px', 'whiteSpace': 'normal'
                            },
                            style_header={
                                'backgroundColor': '#4CAF50',
                                'fontWeight': 'bold',
                                'color': 'white',
                                'border': '1px solid black'
                            },
                            style_data_conditional=[
                                {
                                    'if': {'row_index': 'odd'},
                                    'backgroundColor': '#f2f2f2',
                                }
                            ],
                            tooltip_data=[
                                {
                                    column: {'value': str(value), 'type': 'markdown'}
                                    for column, value in row.items()
                                } for row in fetch_data(r'D:\Downloads\Zulf\Product Details_v2.xlsx', 'prodet').to_dict('records')
                            ],
                            tooltip_duration=None,
                            css=[{
                                'selector': '.dash-cell div.dash-cell-value',
                                'rule': 'display: inline; white-space: inherit; overflow: inherit; text-overflow: inherit;'
                            }]
                        ),
                        dcc.Interval(
                            id='interval-component-table',
                            interval=5000,
                            n_intervals=0
                        )
                    ])
                ]),
                dcc.Tab(label='Similarity Catalogue', value='tab-3', children=[
                html.H2('Below are the similarity details', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                html.Div([
                    dash_table.DataTable(
                        id='similarity-data-table',
                        columns=[{'name': i, 'id': i} for i in fetch_similarity_data().columns],
                        data=fetch_similarity_data().to_dict('records'),
                        filter_action='native',
                        sort_action="native",
                        page_size=10,
                        style_table={'height': '400px', 'overflowY': 'auto', 'marginBottom': '20px'},
                        style_cell={
                            'textAlign': 'center',
                            'padding': '5px',
                            'backgroundColor': '#f9f9f9',
                            'border': '1px solid black',
                            'minWidth': '120px', 'maxWidth': '150px', 'whiteSpace': 'normal'
                        },
                        style_header={
                            'backgroundColor': '#4CAF50',
                            'fontWeight': 'bold',
                            'color': 'white',
                            'border': '1px solid black'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': '#f2f2f2',
                            }
                        ],
                        tooltip_data=[
                            {
                                column: {'value': str(value), 'type': 'markdown'}
                                for column, value in row.items()
                            } for row in fetch_similarity_data().to_dict('records')
                        ],
                        tooltip_duration=None,
                        css=[{
                            'selector': '.dash-cell div.dash-cell-value',
                            'rule': 'display: inline; white-space: inherit; overflow: inherit; text-overflow: inherit;'
                        }]
                    ),
                    dcc.Interval(
                        id='interval-component-similarity-table',
                        interval=5000,
                        n_intervals=0
                    )
                ])
            ]),
                # Modify Tab
                dcc.Tab(label='Modify', value='tab-modify', children=[
                    dcc.Tabs(id='modify-sub-tabs', value='tab-inhouse', children=[
                        dcc.Tab(label='InHouse', value='tab-inhouse', children=[
                            html.Div([
                                html.H2('Modify InHouse Product Data',
                                        style={'textAlign': 'left', 'marginBottom': '20px', 'marginTop': '20px',
                                            'fontSize': '15px'}),

                                dbc.Row([
                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='inhouse-product-dropdown',
                                            placeholder='Select Product Name',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=3
                                    ),

                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='inhouse-component-dropdown',
                                            placeholder='Select Component',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=3
                                    )
                                ]),

                                dbc.Row([
                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='inhouse-column-dropdown',
                                            placeholder='Select field to Edit',
                                            style={'marginBottom': '20px'}
                                        ), width=4
                                    )
                                ]),

                                dbc.Row([
                                    dbc.Col(
                                        html.Div(id='inhouse-value-container'),
                                        width=6
                                    ),
                                    dbc.Col(
                                        html.Button('Confirm Changes', id='inhouse-confirm-changes-button', n_clicks=0,
                                                    style={'marginTop': '20px'}),
                                        width=6
                                    )
                                ]),

                                html.Div(id='inhouse-confirm-message', style={'marginTop': '20px', 'color': 'green',
                                                                            'fontWeight': 'bold'}),

                                # DataTable to display selected data
                                html.Div([
                                    dash_table.DataTable(
                                        id='inhouse-selected-data-table',
                                        columns=[
                                            {'name': 'UniqueID', 'id': 'UniqueID'},
                                            {'name': 'Product Name', 'id': 'Product Name'},
                                            {'name': 'Order Processing Date', 'id': 'Order Processing Date'},
                                            {'name': 'Promised Delivery Date', 'id': 'Promised Delivery Date'},
                                            {'name': 'Quantity Required', 'id': 'Quantity Required'},
                                            {'name': 'Components', 'id': 'Components'},
                                            {'name': 'Operation', 'id': 'Operation'},
                                            {'name': 'Process Type', 'id': 'Process Type'},
                                            {'name': 'Machine Number', 'id': 'Machine Number'},
                                            {'name': 'Processing Time', 'id': 'Run Time (min/1000)'},
                                            {'name': 'Start Time', 'id': 'Start Time'},
                                            {'name': 'End Time', 'id': 'End Time'},
                                            {'name': 'Status', 'id': 'Status'}
                                        ],
                                        data=[],  # Initially empty until products and components are selected
                                        style_table={'height': '400px', 'overflowY': 'auto'},
                                        style_header={
                                            'backgroundColor': 'rgb(230, 230, 230)',
                                            'fontWeight': 'bold'
                                        },
                                        style_cell={
                                            'textAlign': 'left',
                                            'minWidth': '100px',
                                            'maxWidth': '180px',
                                            'whiteSpace': 'normal'
                                        },
                                        style_data_conditional=[
                                            {
                                                'if': {'row_index': 'odd'},
                                                'backgroundColor': 'rgb(248, 248, 248)'
                                            },
                                            {
                                                'if': {'column_id': 'Status', 'filter_query': '{Status} = "Delayed"'},
                                                'backgroundColor': 'tomato',
                                                'color': 'white',
                                                'fontWeight': 'bold'
                                            }
                                        ],
                                        page_size=10,
                                        sort_action='native',
                                        filter_action='native',
                                        column_selectable='single',
                                        row_selectable='single',
                                        selected_columns=[],
                                        selected_rows=[],
                                        editable=True
                                    )
                                ])
                            ])
                        ]),
                        dcc.Tab(label='Outsource', value='tab-outsource', children=[
                            html.Div([
                                html.H2('Modify Outsource Product Data',
                                        style={'textAlign': 'left', 'marginBottom': '20px', 'marginTop': '20px',
                                            'fontSize': '15px'}),

                                dbc.Row([
                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='outsource-product-dropdown',
                                            placeholder='Select Product Name',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=3
                                    ),

                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='outsource-component-dropdown',
                                            placeholder='Select Component',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=3
                                    )
                                ]),

                                dbc.Row([
                                    dbc.Col(
                                        dcc.Dropdown(
                                            id='outsource-column-dropdown',
                                            placeholder='Select field to Edit',
                                            style={'marginBottom': '20px'}
                                        ), width=4
                                    )
                                ]),

                                dbc.Row([
                                    dbc.Col(
                                        dbc.Input(
                                            id='outsource-value-input',
                                            placeholder='Enter New Value',
                                            type='text',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=6
                                    ),
                                    dbc.Col(
                                        html.Button('Confirm Changes', id='outsource-confirm-changes-button', n_clicks=0,
                                                    style={'marginTop': '20px'}),
                                        width=6
                                    )
                                ]),

                                html.Div(id='outsource-confirm-message', style={'marginTop': '20px', 'color': 'green',
                                                                                'fontWeight': 'bold'}),

                                # DataTable to display selected data
                                html.Div([
                                    dash_table.DataTable(
                                        id='outsource-selected-data-table',
                                        columns=[
                                            {'name': 'UniqueID', 'id': 'UniqueID'},
                                            {'name': 'Product Name', 'id': 'Product Name'},
                                            {'name': 'Order Processing Date', 'id': 'Order Processing Date'},
                                            {'name': 'Promised Delivery Date', 'id': 'Promised Delivery Date'},
                                            {'name': 'Quantity Required', 'id': 'Quantity Required'},
                                            {'name': 'Components', 'id': 'Components'},
                                            {'name': 'Operation', 'id': 'Operation'},
                                            {'name': 'Process Type', 'id': 'Process Type'},
                                            {'name': 'Machine Number', 'id': 'Machine Number'},
                                            {'name': 'Processing Time', 'id': 'Run Time (min/1000)'},
                                            {'name': 'Start Time', 'id': 'Start Time'},
                                            {'name': 'End Time', 'id': 'End Time'},
                                            {'name': 'Status', 'id': 'Status'}
                                        ],
                                        data=[],  # Initially empty until products and components are selected
                                        style_table={'height': '400px', 'overflowY': 'auto', 'marginBottom': '20px'},
                                        style_header={
                                            'backgroundColor': 'rgb(230, 230, 230)',
                                            'fontWeight': 'bold'
                                        },
                                        style_cell={
                                            'textAlign': 'left',
                                            'minWidth': '100px',
                                            'maxWidth': '180px',
                                            'whiteSpace': 'normal'
                                        },
                                        style_data_conditional=[
                                            {
                                                'if': {'row_index': 'odd'},
                                                'backgroundColor': 'rgb(248, 248, 248)'
                                            },
                                            {
                                                'if': {'column_id': 'Status', 'filter_query': '{Status} = "Delayed"'},
                                                'backgroundColor': 'tomato',
                                                'color': 'white',
                                                'fontWeight': 'bold'
                                            }
                                        ],
                                        page_size=10,
                                        sort_action='native',
                                        filter_action='native',
                                        column_selectable='single',
                                        row_selectable='single',
                                        selected_columns=[],
                                        selected_rows=[],
                                        editable=True
                                    )
                                ])
                            ])
                        ]),
                        dcc.Tab(label='Time Converter', value='tab-time-converter', children=[
                            html.Div([
                                html.H2('Convert Time to Minutes',
                                        style={'textAlign': 'left', 'marginBottom': '20px', 'marginTop': '20px',
                                            'fontSize': '30px'}),

                                dbc.Row([
                                    dbc.Col(
                                        dbc.RadioItems(
                                            id='time-conversion-type',
                                            options=[
                                                {'label': 'Days to Minutes', 'value': 'days'},
                                                {'label': 'Hours to Minutes', 'value': 'hours'},
                                                {'label': 'Minutes to Days', 'value': 'minutes_to_days'}  # New option added here
                                            ],
                                            value='days',
                                            inline=True,  # Display options inline
                                            style={
                                                'fontSize': '16px',  # Increase font size of labels
                                                'marginBottom': '20px'
                                            },
                                            labelStyle={
                                                'display': 'inline-flex',  # Use inline-flex for better alignment and spacing
                                                'alignItems': 'center',  # Center radio button and label vertically
                                                'marginRight': '30px',  # Space between radio buttons
                                                'flexDirection': 'row',  # Align radio buttons horizontally
                                            }
                                        ),
                                        width=6
                                    )
                                ]),

                                dbc.Row([
                                    dbc.Col(
                                        dbc.Input(
                                            id='time-input',
                                            placeholder='Enter Number of Days or Hours or Minutes',
                                            type='number',
                                            style={'marginBottom': '20px'}
                                        ),
                                        width=3
                                    ),
                                    dbc.Col(
                                        html.Button('Convert', id='convert-button', n_clicks=0, style={'marginTop': '20px'}),
                                        width=2
                                    )
                                ]),

                                html.Div(id='conversion-result', style={'marginTop': '20px', 'fontWeight': 'bold', 'color': 'White'})
                            ])
                        ])
                    ])
                ]),
                
                dcc.Tab(label='Visualize', value='tab-output', children=[
                html.Div(
                    "Select a plot to display:",
                    style={'textAlign': 'center', 'fontSize': '18px', 'marginTop': '20px'}
                ),
                dcc.Dropdown(
                    id='plot-dropdown',
                    options=[
                        {'label': 'Gantt Chart', 'value': 'Gantt Chart'},
                        {'label': 'Gantt Chart(Unschedule)', 'value': 'Gantt Chart(Unschedule)'},
                        {'label': 'Utilization', 'value': 'Utilization'},
                        {'label': 'Time Taken by each Machine', 'value': 'Time Taken by each Machine'},
                        {'label': 'Time taken by each product', 'value': 'Time taken by each product'},
                        {'label': 'Wait Time', 'value': 'Wait Time'},
                        {'label': 'Idle Time', 'value': 'Idle Time'},
                        {'label': 'Product Components Status', 'value': 'Product Components Status'},
                        {'label': 'Remaining Time', 'value': 'Remaining Time'}
                    ],
                    value='Gantt Chart',
                    style={'width': '50%', 'margin': '15px auto'}
                ),
                html.Div(
                    id='graph-container',
                    style={'width': '100%', 'height': '600px'},
                    children=[
                        dcc.Graph(
                            id='main-graph',
                            style={'width': '100%', 'height': '100%', 'marginTop': '30px', 'marginBottom': '30px'}
                        ),
                    ]
                ),
                dcc.Interval(
                    id='interval-component-data',
                    interval=5000,  # Update data and chart every 5 seconds (adjust as needed)
                    n_intervals=0
                )
        ]),
                dcc.Tab(label='Results', value='tab-results', children=[
        dcc.Tabs(id='results-sub-tabs', value='tab-prod', children=[
        
            # Product Tab
            dcc.Tab(label='Product', value='tab-product', children=[
                html.Div([
                    dash_table.DataTable(
                        id='product-table',
                        columns=[{"name": i, "id": i} for i in (product_df.columns if product_df is not None else [])],
                        data=product_df.to_dict('records') if product_df is not None else [],
                        style_table={'overflowX': 'auto'},
                        style_cell={
                            'minWidth': '100px', 'width': '150px', 'maxWidth': '300px',
                            'whiteSpace': 'normal',
                            'textAlign': 'center'
                        },
                        style_header={
                            'backgroundColor': 'rgb(230, 230, 230)',
                            'fontWeight': 'bold'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': 'rgb(248, 248, 248)'
                            },
                        ],
                        page_size=10,  # Adjust as needed
                    ) if product_df is not None else html.Div("No data available.", style={'textAlign': 'center'})
                ])
            ]),
            
            # Machine Utilization Tab
            dcc.Tab(label='Machine Utilization', value='tab-machine', children=[
                html.Div([
                    dash_table.DataTable(
                        id='machine-table',
                        columns=[{"name": i, "id": i} for i in (machine_df.columns if machine_df is not None else [])],
                        data=machine_df.to_dict('records') if machine_df is not None else [],
                        style_table={'overflowX': 'auto'},
                        style_cell={
                            'minWidth': '100px', 'width': '150px', 'maxWidth': '300px',
                            'whiteSpace': 'normal',
                            'textAlign': 'center'
                        },
                        style_header={
                            'backgroundColor': 'rgb(230, 230, 230)',
                            'fontWeight': 'bold'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': 'rgb(248, 248, 248)'
                            },
                        ],
                        page_size=10,  # Adjust as needed
                    ) if machine_df is not None else html.Div("No data available.", style={'textAlign': 'center'})
                ])
            ]),
            # Order Details Tab
            dcc.Tab(label='Product Details', value='tab-component', children=[
                html.Div([
                    dash_table.DataTable(
                        id='component-table',
                        columns=[{"name": i, "id": i} for i in (component_df.columns if component_df is not None else [])],
                        data=component_df.to_dict('records') if component_df is not None else [],
                        style_table={'overflowX': 'auto'},
                        style_cell={
                            'minWidth': '100px', 'width': '150px', 'maxWidth': '300px',
                            'whiteSpace': 'normal',
                            'textAlign': 'center'
                        },
                        style_header={
                            'backgroundColor': '#f2f2f2',
                            'fontWeight': 'bold',
                            'border': '1px solid black'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': 'rgb(248, 248, 248)'
                            },
                        ],
                        page_size=10,  # Adjust as needed
                    ) if component_df is not None else html.Div("No data available.", style={'textAlign': 'center'})
                ])
            ]),
            # Order Details Tab
            dcc.Tab(label='Scheduling Details', value='tab-order-details', children=[
                html.Div([
                    dash_table.DataTable(
                        id='order-details-table',
                        columns=[{"name": i, "id": i} for i in (order_details_df.columns if order_details_df is not None else [])],
                        data=order_details_df.to_dict('records') if order_details_df is not None else [],
                        style_table={'overflowX': 'auto'},
                        style_cell={
                            'minWidth': '100px', 'width': '150px', 'maxWidth': '300px',
                            'whiteSpace': 'normal',
                            'textAlign': 'center'
                        },
                        style_header={
                            'backgroundColor': '#f2f2f2',
                            'fontWeight': 'bold',
                            'border': '1px solid black'
                        },
                        style_data_conditional=[
                            {
                                'if': {'row_index': 'odd'},
                                'backgroundColor': 'rgb(248, 248, 248)'
                            },
                        ],
                        page_size=10,  # Adjust as needed
                    ) if order_details_df is not None else html.Div("No data available.", style={'textAlign': 'center'})
                ])
            ]),
        ]),
    ]),
        # Instructions section
        dcc.Tab(label='Instructions', value='tab-instructions', children=[
            html.Div([
                html.H2('Instructions', style={'marginTop': '30px'}),
                html.Ul([  # Use Unordered List for the points
                    html.Li("The Read from spreadsheet button, checks if the Product Details file is valid and retrieves details"),
                    html.Li("The Start button, starts the scheduling of products with data in Product details file."),
                    html.Li("The Pause button, used to pause the process for making any changes in the data"),
                    html.Li("The Reschedule button is used to continue the scheduling after the changes are made"),
                    html.Li("The Reset button, reset the data to intial status"),
                    html.Li("Use the 'Product List change' tab to add, delete, or swap products in your list."),
                    html.Li("The 'Product Catalogue' tab shows detailed product information of the Products/components that are Inprogress or Yet to start. You can filter and sort the data."),
                    html.Li("'Similarity Catalogue' provides details on the similarity analysis between products."),
                    html.Li("In the 'Modify' tab, you can adjust product data for both in-house and outsourced products."),
                    html.Li("The 'Visualise' tab contains, Gantt chart, KPI's visualization"),
                    html.Li("The 'Result' tab provides the priority of products and Product Details"),
                    html.Li("The dashboard refreshes automatically. Use the buttons above to Read spreadsheet, start, pause, or restart after updating data.")
                ], style={'fontSize': '16px', 'lineHeight': '1.6em'})
            ])
        ])


            ], style={'width': '100%','marginTop': '50px', 'marginBottom': '50px'}),
        
        ]
    )

    @app.callback(
    Output('completion-message', 'children'),
    Output('completion-message', 'className'),
    Input('interval-component-data', 'n_intervals')
    )
    def update_completion_message(n_intervals):
        # Fetch data from your source
        df1 = fetch_data_from_excel(full_file_path, 'prodet')
        #df1 = product_df  # Example data for testing
        if df1 is None or df1.empty:
            raise ValueError("Data is empty or not loaded correctly.")
        statuses = set(df1['Status'])
        if statuses.issubset({'Completed', 'Late'}):
            return "Production scheduling is Completed", 'message-box highlight-green'
        else:
            return "Production Scheduling is InProgress", 'message-box blinking'
        
    @app.callback(
    Output('conversion-result', 'children'),
    Input('convert-button', 'n_clicks'),
    State('time-conversion-type', 'value'),
    State('time-input', 'value')
    )
    def convert_time(n_clicks, conversion_type, time_value):
        if n_clicks > 0 and time_value is not None:
            if conversion_type == 'days':
                minutes = round(time_value * 24 * 60, 2)  # Convert days to minutes and round to 2 decimal points
                return f'{time_value} day(s) is equal to {minutes} minute(s).'
            elif conversion_type == 'hours':
                minutes = round(time_value * 60, 2)  # Convert hours to minutes and round to 2 decimal points
                return f'{time_value} hour(s) is equal to {minutes} minute(s).'
            elif conversion_type == 'minutes_to_days':
                days = round(time_value / (24 * 60), 2)  # Convert minutes to days
                return f'{time_value} minute(s) is equal to {days} day(s).'
        return "Enter a valid number and click Convert."

    @app.callback(
        Output('manage-content', 'children'),
        Input('manage-dropdown', 'value')
    )
    def render_manage_content(action):
        if action == 'add':
            return html.Div(
                id='input-form',
                children=[
                    html.H2('Add New Product', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Sr. No:"),
                            dbc.Input(id='Sr-No', type='number', placeholder='Enter product number (Product 1, Product 2, ...)'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Product Name:"),
                            dbc.Input(id='Product-Name', type='text', placeholder='Enter product name'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Order Processing Date:"),
                            dbc.Input(id='Order-Processing-Date', type='date', placeholder='Enter processing date'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Promised Delivery Date:"),
                            dbc.Input(id='Promised-Delivery-Date', type='date', placeholder='Enter delivery date'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Quantity Required:"),
                            dbc.Input(id='Quantity-Required', type='number', placeholder='Enter required quantity'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Components:"),
                            dbc.Input(id='Components', type='text', placeholder='Enter components'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Operation:"),
                            dbc.Input(id='Operation', type='text', placeholder='Enter operation'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                                [
                                    dbc.InputGroupText("Process Type:"),
                                    dcc.Dropdown(
                                        id='Process-Type',
                                        options=[
                                            {'label': 'In House', 'value': 'In House'},
                                            {'label': 'Outsource', 'value': 'Outsource'}
                                        ],
                                        placeholder='Select process type...',
                                        style={'width': '70%'}  # Adjust width as needed
                                    ),
                                ],
                                style={'marginBottom': '10px'}
                            ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Machine Number:"),
                            dbc.Input(id='Machine-Number', type='text', placeholder='Enter machine number'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Run Time (min/1000):"),
                            dbc.Input(id='Run-Time', type='number', placeholder='Enter run time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Cycle Time (seconds):"),
                            dbc.Input(id='Cycle-Time', type='number', placeholder='Enter cycle time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Setup time (seconds):"),
                            dbc.Input(id='Setup-Time', type='number', placeholder='Enter setup time'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    html.Button('Submit', id='submit-button', n_clicks=0, style={'marginTop': '10px'}),
                    html.Div(id='submit-output', style={'marginTop': '10px'})
                ]
            )

        elif action == 'delete':
            return html.Div(
                id='input-form',
                children=[
                    html.H2('Delete Product', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("UniqueID:"),
                            dbc.Input(id='UniqueID-delete', type='number', placeholder='Enter UniqueID to delete'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    html.Button('Delete', id='delete-button', n_clicks=0, style={'marginTop': '10px'}),
                    html.Div(id='delete-output', style={'marginTop': '10px'})
                ]
            )

        elif action == 'swap':
            return html.Div(
                id='input-form',
                children=[
                    html.H2('Swap Product', style={'textAlign': 'left', 'marginBottom': '30px', 'fontSize': '20px'}),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("First UniqueID:"),
                            dbc.Input(id='UniqueID-swap1', type='number', placeholder='Enter first UniqueID'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    dbc.InputGroup(
                        [
                            dbc.InputGroupText("Second UniqueID:"),
                            dbc.Input(id='UniqueID-swap2', type='number', placeholder='Enter second UniqueID'),
                        ],
                        style={'marginBottom': '10px'}
                    ),
                    html.Button('Swap', id='swap-button', n_clicks=0, style={'marginTop': '10px'}),
                    html.Div(id='swap-output', style={'marginTop': '10px'})
                ]
            )

        else:
            return html.Div()

    def modify_new_data(new_data, full_file_path):
        try:
            # Load the Excel file
            print(full_file_path)
            wb = load_workbook(full_file_path)
            sheet = wb["Addln"]
            
            # Find the next available row for insertion
            next_row = sheet.max_row + 1
            print(f"Next row{next_row}")
            # Assuming new_data is a dictionary or tuple containing the values
            sheet.append(new_data)
            
            # Save the changes
            wb.save(full_file_path)
            wb.close()
            
            return True
        
        except Exception as e:
            print(f"Error modifying data: {e}")
            return False
        
    # Callback to add new product
    @app.callback(
        Output('submit-output', 'children'),
        Input('submit-button', 'n_clicks'),
        State('Sr-No', 'value'),
        State('Product-Name', 'value'),
        State('Order-Processing-Date', 'value'),
        State('Promised-Delivery-Date', 'value'),
        State('Quantity-Required', 'value'),
        State('Components', 'value'),
        State('Operation', 'value'),
        State('Process-Type', 'value'),
        State('Machine-Number', 'value'),
        State('Run-Time', 'value'),
        State('Cycle-Time', 'value'),
        State('Setup-Time', 'value')
    )
    def add_new_product(n_clicks, sr_no, product_name, processing_date, delivery_date, quantity_required,
                        components, operation, process_type, machine_number, run_time, cycle_time, setup_time):
        if n_clicks > 0:
            try:
                # Load the Excel workbook
                wb = load_workbook(full_file_path)
            
            # Check if the sheet exists in the workbook
                if "Addln" in wb.sheetnames:
                    sheet = wb["Addln"]
                    addln_count = sheet.max_row
                print(f"Addln count{addln_count}")

                if addln_count == 1:
                    # If "Addln" table is empty, get last ID from "prodet"
                    last_id = get_last_unique_id("prodet",full_file_path)
                    print(f"LAst Unique ID ID{last_id}")
                else:
                    # Otherwise, get last ID from "Addln"
                    last_id = get_last_unique_id("Addln",full_file_path)
                print(f"LAst Unique ID{last_id}")
                new_id = last_id + 1  # Increment the last ID by 1
                print(f"New ID{new_id}")
                # Create a tuple or list with the data to be inserted
                new_data = (new_id, sr_no, product_name, processing_date, delivery_date, quantity_required,
                            components, operation, process_type, machine_number, run_time, cycle_time, setup_time)
                
                success = modify_new_data(new_data, full_file_path)
                
                if success:
                    return dbc.Alert("Product added successfully", color="success", dismissable=True)
                else:
                    return dbc.Alert("Error adding product", color="danger", dismissable=True)

                r
            except Exception as e:
                return dbc.Alert(f"Error adding product: {e}", color="danger", dismissable=True)

        return html.Div()


    # Callback to delete product
    @app.callback(
        Output('delete-output', 'children'),
        Input('delete-button', 'n_clicks'),
        State('UniqueID-delete', 'value')
    )
    def delete_product(n_clicks, unique_id):
        if n_clicks > 0:
            try:
                # Load the workbook
                wb = load_workbook(full_file_path)
                sheet = wb["prodet"]

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                    cell = row[0]  # Assuming UniqueID is in the first column
                    if cell.value == unique_id:
                        sheet.delete_rows(cell.row)
                        break

                # Save the changes
                wb.save(full_file_path)
                wb.close()

                return html.Div("Product deleted successfully", style={'color': 'green'})

                

            except Exception as e:
                return dbc.Alert(f"Error deleting product: {e}", color="danger", dismissable=True)

        return html.Div()


    # Callback to swap products
    @app.callback(
        Output('swap-output', 'children'),
        Input('swap-button', 'n_clicks'),
        State('UniqueID-swap1', 'value'),
        State('UniqueID-swap2', 'value')
    )
    def swap_products(n_clicks, unique_id1, unique_id2):
        if n_clicks > 0:
            try:
                # Load the workbook
                wb = load_workbook(full_file_path)
                sheet = wb["prodet"]

                row1_index = None
                row2_index = None
                row1_data = None
                row2_data = None

                # Find the rows with the matching UniqueIDs and store their indices and data
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                    if row[0] == unique_id1:
                        row1_index = row[0]
                        row1_data = list(row[1:])  # Exclude UniqueID and convert tuple to list
                    elif row[0] == unique_id2:
                        row2_index = row[0]
                        row2_data = list(row[1:])  # Exclude UniqueID and convert tuple to list

                # Check if both UniqueIDs were found
                if row1_index is not None and row2_index is not None:
                    # Swap data in the rows
                    for col_idx, value in enumerate(row1_data, start=1):
                        sheet.cell(row=row1_index, column=col_idx+1).value = row2_data[col_idx-1]
                        sheet.cell(row=row2_index, column=col_idx+1).value = value

                    # Save the changes
                    wb.save(full_file_path)
                    wb.close()
                    return f'Products with Unique IDs {unique_id1} and {unique_id2} swapped successfully!'
                else:
                    return 'One or both products not found!'

            except Exception as e:
                return dbc.Alert(f"Error swapping products: {e}", color="danger", dismissable=True)

        return html.Div()


    def fetch_data_Details1(file_path, sheet_name, usecols=None):
        try:
            data = pd.read_excel(file_path, sheet_name=sheet_name, usecols=usecols)
            return data
        except Exception as e:
            print(f"Error fetching data from {sheet_name}: {e}")
            return None
        
    @app.callback(
        Output('data-table', 'columns'),
        Output('data-table', 'data'),
        Input('interval-component-table', 'n_intervals')
    )
    def update_table(n_intervals):
        prodet_columns = ['UniqueID','Sr. No','Product Name','Order Processing Date','Promised Delivery Date','Quantity Required','Components','Operation','Process Type','Machine Number',"Run Time (min/1000)","Start Time","End Time",'Status']
        df = fetch_data_Details1(full_file_path, 'prodet', usecols=prodet_columns)
        # Filter the data to include only rows where 'Status' is not 'Completed'
        filtered_data = df[df['Status'] != 'Completed']
        
        
        if filtered_data is not None:
            columns = [{'name': col, 'id': col} for col in filtered_data.columns]
            data = filtered_data.to_dict('records')
        else:
            columns = []
            data = []
            
        return columns, data



    def  fetch_data_Details(product_name, process_type, component_name=None):
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            print(product_name)
            print(component_name)
            # Apply filters if provided
            if product_name:
                df = df[df['Product Name'] == product_name]
            df=df[df['Process Type']==process_type]
            df=df[df['Status']!="Completed"]
            if component_name:
                df = df[df['Components'] == component_name]
            print(df)
            wb.close()
            return df
        except Exception as e:
            print(f"Error fetching data: {e}")
            return pd.DataFrame()

    def  fetch_data_Details2(product_name, component_name):
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            print(product_name)
            print(component_name)
            # Apply filters if provided
            if product_name:
                df = df[df['Product Name'] == product_name]
            
            if component_name:
                df = df[df['Components'] == component_name]
            print(df)
            wb.close()
            return df
        except Exception as e:
            print(f"Error fetching data: {e}")
            return pd.DataFrame()


    # Sample function to update database
    def modify_DB(unique_id, column_name, new_value):
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            
            # Find the row with the matching unique ID and update the specified column
            df.loc[df['UniqueID'] == unique_id, column_name] = new_value
            
            # Write the DataFrame back to the sheet
            for row in sheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    col_index = cell.column - 1
                    if col_index < len(df.columns):
                        cell.value = df.iloc[cell.row-2, col_index]
            
            wb.save(full_file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Error updating database: {e}")
            return False

    def fetch_products_and_components(process_type):
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            
            # Filter based on process type
            df_filtered = df[df['Process Type'] == process_type]
            print(df_filtered)
            
            # Get distinct product names and components
            product_names = df_filtered['Product Name'].unique().tolist()
            components = df_filtered['Components'].unique().tolist()

            wb.close()
            return product_names, components
        except Exception as e:
            print(f"Error fetching products and components: {e}")
            return [], []


    def fetch_unique_id(product_name, component, process_type):
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            
            # Find the row with the matching criteria
            df_filtered = df[(df['Product Name'] == product_name) & (df['Components'] == component) & (df['Process Type'] == process_type)]
            
            if not df_filtered.empty:
                unique_id = df_filtered.iloc[0]['UniqueID']
                wb.close()
                return unique_id
            else:
                wb.close()
                return None
        except Exception as e:
            print(f"Error fetching UniqueID: {e}")
            return None




    # Callback to render content based on tab selection
    @app.callback(
        Output('tabs-content', 'children'),
        Input('modify-sub-tabs', 'value')
    )
    def render_tab_content(tab):
        if tab == 'tab-inhouse':
            return html.Div([
                dcc.Dropdown(id='inhouse-product-dropdown', placeholder='Select Product'),
                dcc.Dropdown(id='inhouse-component-dropdown', placeholder='Select Component'),
                dcc.Dropdown(id='inhouse-column-dropdown', placeholder='Select Column'),
                html.Div(id='inhouse-value-container'),
                html.Button('Confirm Changes', id='inhouse-confirm-changes-button'),
                html.Div(id='inhouse-confirm-message'),
                dash_table.DataTable(id='inhouse-selected-data-table', columns=[{'name': col, 'id': col} for col in ['Product Name', 'Components', 'Order Processing Date', 'Promised Delivery Date']], data=[]),
            ])
        elif tab == 'tab-outsource':
            return html.Div([
                dcc.Dropdown(id='outsource-product-dropdown', placeholder='Select Product'),
                dcc.Dropdown(id='outsource-component-dropdown', placeholder='Select Component'),
                dcc.Dropdown(id='outsource-column-dropdown', placeholder='Select Column'),
                dbc.Input(id='outsource-value-input', placeholder='Enter New Value', type='text'),
                html.Button('Confirm Changes', id='outsource-confirm-changes-button'),
                html.Div(id='outsource-confirm-message'),
                dash_table.DataTable(id='outsource-selected-data-table', columns=[{'name': col, 'id': col} for col in ['Product Name', 'Components', 'Run Time (min/1000)']], data=[]),
            ])
        else:
            return html.Div()

    # Callback to update product dropdown options for InHouse
    @app.callback(
        Output('inhouse-product-dropdown', 'options'),
        Input('modify-sub-tabs', 'value')
    )
    def update_inhouse_product_dropdown(tab):
        if tab != 'tab-inhouse':
            raise dash.exceptions.PreventUpdate
        
        product_names, _ = fetch_products_and_components('In House')
        
        options = [{'label': name, 'value': name} for name in product_names]
        
        return options

    # Callback to update component dropdown options for InHouse based on selected product
    @app.callback(
        Output('inhouse-component-dropdown', 'options'),
        Input('inhouse-product-dropdown', 'value')
    )
    def update_inhouse_component_dropdown(product_name):
        if product_name is None:
            raise PreventUpdate
        
        df = fetch_data_Details(product_name,"In House")
        components = df['Components'].unique().tolist()

        print(components)
        
        options = [{'label': comp, 'value': comp} for comp in components]
        
        return options

    # Callback to update column dropdown options for InHouse
    @app.callback(
        Output('inhouse-column-dropdown', 'options'),
        Input('inhouse-product-dropdown', 'value'),
        Input('inhouse-component-dropdown', 'value')
    )
    def update_inhouse_column_dropdown(product_name, component_name):
        if product_name is None or component_name is None:
            raise dash.exceptions.PreventUpdate
        
        columns = ['Product Name', 'Order Processing Date', 'Promised Delivery Date', 'Quantity Required', 'Components', 'Operation', 'Process Type', 'Machine Number', 'Run Time (min/1000)', 'Cycle Time (seconds)', 'Setup time (seconds)']
        
        return [{'label': col, 'value': col} for col in columns]

    # Callback to render input type based on selected column in InHouse tab
    @app.callback(
        Output('inhouse-value-container', 'children'),
        Input('inhouse-column-dropdown', 'value')
    )
    def render_inhouse_value_input(column_selected):
        if column_selected is None:
            return html.Div()
        if column_selected in ['Order Processing Date', 'Promised Delivery Date']:
            return html.Div([
                dcc.DatePickerSingle(
                    id={'type': 'dynamic-date-picker', 'index': column_selected + '-date'},
                    display_format='YYYY-MM-DD',
                    style={'marginBottom': '20px'}
                ),
                dcc.Input(
                    id={'type': 'dynamic-time-picker', 'index': column_selected + '-time'},
                    type='text',
                    placeholder='Select Time (HH:MM)',
                    style={'marginBottom': '20px'}
                )
            ])
        else:
            return dbc.Input(
                id={'type': 'dynamic-input', 'index': column_selected},
                placeholder='Enter New Value',
                type='text',
                style={'marginBottom': '20px'}
            )
    def get_time_from_input(column_name,time_str):
        try:
            print(time_str)
            hours, minutes = map(int, time_str.split(':'))
            return time(hours, minutes)
        except ValueError:
            return None


    # Callback to handle confirmation of changes in InHouse tab
    @app.callback(
        Output('inhouse-confirm-message', 'children'),
        Input('inhouse-confirm-changes-button', 'n_clicks'),
        State('inhouse-product-dropdown', 'value'),
        State('inhouse-component-dropdown', 'value'),
        State('inhouse-column-dropdown', 'value'),
        State({'type': 'dynamic-input', 'index': ALL}, 'value'),
        State({'type': 'dynamic-date-picker', 'index': ALL}, 'date'),
        State({'type': 'dynamic-time-picker', 'index': ALL}, 'value'),
    )
    def update_inhouse_database(n_clicks, product_name, component_name, column_name, new_values, date_values,time_values):
        try: 
            if n_clicks == 0 or product_name is None or component_name is None or column_name is None:
                raise dash.exceptions.PreventUpdate
            print(f"new_values: {new_values}")
            print(f"date_values: {date_values}")
            print(f"Column Selected: {column_name}")
            print(f"Time: {time_values}")
            if not all(new_values) and not all(date_values):
                return "Please fill all fields."
            
            #db_name = 'ProductDetails'
            #db_username = 'PUser12'
            #db_password = 'PSQL@123'
            #db_host = 'localhost'
            #db_port = '5432'
            
            unique_id = fetch_unique_id(product_name, component_name, 'In House')
            #print(unique_id)
            success_messages = []
            #print(new_values)
            
            for idx, value in enumerate(new_values):
                print(f"value: {value}")
                if value:
                    success = modify_DB(unique_id, column_name, value)
                    if success:
                        success_messages.append(f"Successfully updated {column_name} to {value}.")
            for idx, date_val in enumerate(date_values):
                if date_val:
                    #time_val = get_time_from_input(column_name,time_values)  # Extract time from input
                    time_val=time_values[0]
                    print(f"Date Value: {date_val}")
                    print(f"Time Value: {time_val}")
                    if time_val:
                        #datetime_val = datetime.combine(date_val, time_val)
                        datetime_str = f'{date_val}T{time_val}'
                        print(datetime_str)
                        success = modify_DB(unique_id, column_name, datetime_str)
                        if success:
                            success_messages.append(f"Successfully updated {column_name} to {datetime_str}.")
                    else:
                        success = modify_DB(unique_id, column_name, date_val)
                        if success:
                            success_messages.append(f"Successfully updated {column_name} to {date_val}.")
            
            if success_messages:
                return html.Div([html.P(msg) for msg in success_messages])
            else:
                return "Error updating database. Please try again."
        except Exception as e:
            print(f"Error in update_inhouse_database: {e}")
    # Callback to update product dropdown options for Outsource
    @app.callback(
        Output('outsource-product-dropdown', 'options'),
        Input('modify-sub-tabs', 'value')
    )
    def update_outsource_product_dropdown(tab):
        if tab != 'tab-outsource':
            raise dash.exceptions.PreventUpdate
        
        product_names, _ = fetch_products_and_components('Outsource')
        
        options = [{'label': name, 'value': name} for name in product_names]
        
        return options

    # Callback to update component dropdown options for Outsource based on selected product
    @app.callback(
        Output('outsource-component-dropdown', 'options'),
        Input('outsource-product-dropdown', 'value')
    )
    def update_outsource_component_dropdown(product_name):
        if product_name is None:
            raise PreventUpdate
        df = fetch_data_Details(product_name,"Outsource")
        components = df['Components'].unique().tolist()
        
        options = [{'label': comp, 'value': comp} for comp in components]
        
        return options

    # Callback to update column dropdown options for Outsource
    @app.callback(
        Output('outsource-column-dropdown', 'options'),
        Input('outsource-product-dropdown', 'value'),
        Input('outsource-component-dropdown', 'value')
    )
    def update_outsource_column_dropdown(product_name, component_name):
        if product_name is None or component_name is None:
            raise dash.exceptions.PreventUpdate
        
        columns =  columns = {
        'Run Time (min/1000)': 'Processing Time'
    }
        
        # Return options with display names
        return [{'label': display_name, 'value': col} for col, display_name in columns.items()]

    # Callback to handle confirmation of changes in Outsource tab
    @app.callback(
        Output('outsource-confirm-message', 'children'),
        Input('outsource-confirm-changes-button', 'n_clicks'),
        State('outsource-product-dropdown', 'value'),
        State('outsource-component-dropdown', 'value'),
        State('outsource-column-dropdown', 'value'),
        State('outsource-value-input', 'value')
    )
    def update_outsource_database(n_clicks, product_name, component_name, column_name, new_value):
        if n_clicks == 0 or product_name is None or component_name is None or column_name is None:
            raise dash.exceptions.PreventUpdate
        
        if new_value is None:
            return "Please fill all fields."
              
        unique_id = fetch_unique_id(product_name, component_name, 'Outsource')
        success = modify_DB(unique_id, column_name, new_value)
        
        if success:
            return f"Successfully updated {column_name} to {new_value}."
        else:
            return "Error updating database. Please try again."

    # Callback to update the data table based on product and component selection for InHouse
    @app.callback(
        Output('inhouse-selected-data-table', 'data'),
        Input('inhouse-product-dropdown', 'value'),
        Input('inhouse-component-dropdown', 'value')
    )
    def update_inhouse_selected_data_table(product_name, component_name):
        if product_name and component_name:
            df = fetch_data_Details2(product_name, component_name)
            return df.to_dict('records')
        else:
            return []

    # Callback to update the data table based on product and component selection for Outsource
    @app.callback(
        Output('outsource-selected-data-table', 'data'),
        Input('outsource-product-dropdown', 'value'),
        Input('outsource-component-dropdown', 'value')
    )
    def update_outsource_selected_data_table(product_name, component_name):
        if product_name and component_name:
            df = fetch_data_Details2(product_name, component_name)
            return df.to_dict('records')
        else:
            return []


    #allocation_process=None
    # Function to read the stdout and stderr
    def read_output(process):
        try:
            for stdout_line in iter(process.stdout.readline, b''):
                if stdout_line:
                    print(stdout_line.decode().strip())
            process.stdout.close()
        except Exception as e:
            print(f"Error reading stdout: {e}")

        try:
            for stderr_line in iter(process.stderr.readline, b''):
                if stderr_line:
                    print(f"ERROR: {stderr_line.decode().strip()}")
            process.stderr.close()
        except Exception as e:
            print(f"Error reading stderr: {e}")

    def start_allocation_check(keyword):
        global allocation_process
        # Print statements to debug
        print("Starting Allocation_check_Excel.py...")
        print(f"Keyword: {keyword}")
        
        # Verifying if the file exists
        script_path = r'Allocation_check_Excel.py'
        if not os.path.exists(script_path):
            print(f"Script not found at path: {script_path}")
            return
        try:
        
            allocation_process = subprocess.Popen(
                [sys.executable, script_path],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1  # Line-buffered mode
            )
            #allocation_process.stdin.write(keyword + '\n')
            #allocation_process.stdin.flush()
            stdout, stderr = allocation_process.communicate(keyword)
            #print("Script output:", stdout)
            #print("Script error:", stderr)
            threading.Thread(target=read_output, args=(allocation_process,), daemon=True).start()
            print("Allocation_check.py started successfully")
        except Exception as e:
            print(f"Error starting Allocation_check.py: {e}")
            raise
            


    # Function to stop the execution of the external script
    def stop_allocation_check():
        global allocation_process
        if allocation_process and allocation_process.poll() is None:
            allocation_process.terminate()
            allocation_process = None
    
    @app.callback(
        [
            Output('interval-component-script', 'disabled'),
            Output('start-message', 'children'),
            Output('read-modal', 'is_open'),
            Output('initialise-modal', 'is_open'),
            Output('start-modal', 'is_open'),
            Output('stop-modal', 'is_open'),
            Output('reset-modal', 'is_open'),
            Output('read-modal-body', 'children'),
            Output('initialise-modal-body', 'children'),
            Output('start-modal-body', 'children'),
            Output('stop-modal-body', 'children'),
            Output('reset-modal-body', 'children')
        ],
        [
            Input('read-button', 'n_clicks'),
            Input('initialise-button', 'n_clicks'),
            Input('start-button', 'n_clicks'),
            Input('stop-button', 'n_clicks'),
            Input('reset-button', 'n_clicks'),
            Input('close-read-modal', 'n_clicks'),
            Input('close-initialise-modal', 'n_clicks'),
            Input('close-start-modal', 'n_clicks'),
            Input('close-stop-modal', 'n_clicks'),
            Input('close-reset-modal', 'n_clicks')
        ],
        [State('interval-component-script', 'disabled')]
    )
    def control_allocation_check(read_clicks, initialise_clicks, start_clicks, stop_clicks, reset_clicks, close_read_modal,
                                close_initialise_modal, close_start_modal, close_stop_modal, close_reset_modal,
                                interval_disabled):
        ctx = dash.callback_context
        if not ctx.triggered:
            return interval_disabled, None, False, False, False, False, False, "", "", "", "", ""
        
        triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]
        read_modal_open=False
        initialise_modal_open = False
        start_modal_open = False
        stop_modal_open = False
        reset_modal_open = False
        read_modal_body = ""
        initialise_modal_body = ""
        start_modal_body = ""
        stop_modal_body = ""
        modal_body = "Process started..."

        if triggered_id == 'read-button' and read_clicks:
            # Open the modal immediately
            read_modal_open = True
           
            if os.path.exists(full_file_path):
                print("File Exists")
                modal_body = "Data is retrieved from spreadsheet. Click on the Start button to start scheduling process"
            else:
                print("File Not found, Please check")
                modal_body = "Product Details File Not found, Please check"
            # Start the process asynchronously
            #threading.Thread(target=start_allocation_check, args=("Initial",), daemon=True).start()
            interval_disabled = False
            
        elif triggered_id == 'initialise-button' and initialise_clicks:
            # Open the modal immediately
            initialise_modal_open = True
            modal_body = "Scheduling Process Started."
            # Start the process asynchronously
            threading.Thread(target=start_allocation_check, args=("Initial",), daemon=True).start()
            interval_disabled = False
 
        elif triggered_id == 'start-button' and start_clicks:
            start_modal_open = True
            modal_body = "Scheduling Process Started again."
            threading.Thread(target=start_allocation_check, args=("Start",), daemon=True).start()
            interval_disabled = False

        elif triggered_id == 'stop-button' and stop_clicks:
            stop_modal_open = True
            modal_body = "Program paused, Make the necessary changes, If any. and click on the Reschedule button."
            stop_allocation_check()
            interval_disabled = True
        elif triggered_id == 'reset-button' and reset_clicks:
            reset_modal_open = True
            modal_body = "Product Details have been reset."

            # Perform the reset functionality
            if os.path.exists(full_file_path):
                try:
                    # Read the 'P' sheet
                    df_p = pd.read_excel(full_file_path, sheet_name='P')
                    
                    # Load the entire Excel file to write back to the same file
                    with pd.ExcelWriter(full_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # Update the 'prodet' sheet
                        df_p.to_excel(writer, sheet_name='prodet', index=False)
                        
                    modal_body += " Data copied from 'P' sheet to 'prodet' sheet successfully."
                except Exception as e:
                    modal_body = f"An error occurred: {e}"
            else:
                modal_body = "Product Details File Not found, Please check"
            
            interval_disabled = True

        elif triggered_id.startswith('close'):
            return interval_disabled, None, False, False, False, False, False, "", "", "","",""
        
        return (interval_disabled, None,read_modal_open, initialise_modal_open, start_modal_open, stop_modal_open, reset_modal_open,
                modal_body if read_modal_open else "",
                modal_body if initialise_modal_open else "",
                modal_body if start_modal_open else "",
                modal_body if stop_modal_open else "",
                modal_body if reset_modal_open else "")


    def fetch_latest_completed_time():
        try:
            wb = load_workbook(full_file_path)
            sheet = wb["prodet"]
            
            # Read the sheet into a DataFrame
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)
            
            # Filter the DataFrame for completed statuses and get the latest end time
            df_completed = df[df['Status'] == 'Completed']
            if df_completed.empty:
                return None
            latest_end_time = df_completed['End Time'].max()
            
            wb.close()
            return latest_end_time
        except Exception as e:
            print(f"Error fetching latest completed time: {e}")
            return None
    # Function to fetch data from the database
    def fetch_data_runtime():
        try:
            wb = load_workbook('RunTime.xlsx')
            sheet = wb["Sheet1"]
            
            # Assuming 'RunTime' data is in the first row, first column
            run_time = sheet.cell(row=2, column=1).value
            
            if run_time is None:
                run_time1 = 0  # Default value if no run_time is found
            else:
                run_time1 = float(run_time)
            #print(f"Run Time{run_time1}")
            wb.close()
            return run_time1
        except Exception as e:
            print(f"Error fetching data: {e}")
            return 0
            
    # Assume this is your global start_time initialized somewhere in your app
    #Dash_time = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)  # Example start time
    #Dash_time = datetime.now()
    # Function to fetch data from the database
    def update_data_runtime():
        try:
            conn = psycopg2.connect(
                dbname=db_name,
                user=db_username,
                password=db_password,
                host=db_host,
                port=db_port
            )
            # Set the runtime data to NULL
            update_query = '''UPDATE public."RunTime" SET "Run_time" = NULL;'''
                
            with conn.cursor() as cursor:
                cursor.execute(update_query)
                
                
                
                conn.commit()

            conn.close()
            
            
        except Exception as e:
            print(f"Error updating data: {e}")
            
    # Callback to update the live clock and date
    @app.callback(
        Output('live-clock', 'children'),
        Output('current-date', 'children'),
        Output('current-day', 'children'),
        Input('interval-component-clock', 'n_intervals')
    )
    def update_clock(n):
        global Dash_time
        
        Dash_time = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)
        

        current_time = Dash_time.strftime('%H:%M:%S')
        current_date = Dash_time.strftime('%d-%m-%Y')
        current_day = Dash_time.strftime('%A')
        
        return current_time, current_date, current_day



    @app.callback(
        Output('main-graph', 'figure'),
        [Input('plot-dropdown', 'value'), Input('interval-component-data', 'n_intervals')]
    )
    def update_graph(selected_plot,n):
        #df = fetch_data()
        
        df = fetch_data(full_file_path, "prodet")
        
        if df.empty:
            return px.line(title='No Data Available')

        # Process the DataFrame for each plot type
        if selected_plot=="Gantt Chart":
            
            # Convert 'Start Time' and 'End Time' columns to datetime, handling non-datetime values
            df['Start Time'] = pd.to_datetime(df['Start Time'], errors='coerce')
            df['End Time'] = pd.to_datetime(df['End Time'], errors='coerce')

            # Create new DataFrame for plotting
            plot_data = []

            # Group by Product Name to process each product individually
            for product_name, product_group in df.groupby('Product Name'):
                total_delay = timedelta(0)  # Initialize total delay for each product
                delay_count = 0  # Counter to track how many delays exist
                last_end_time = None
                last_machine_number = ''  # Variable to store the machine number of the last component

                for index, row in product_group.iterrows():
                    # Add a small black line if Process Type is 'In House'
                    if row['Process Type'] == 'In House':
                        setup_start = row['Start Time'] - timedelta(minutes=30) if pd.notnull(row['Start Time']) else pd.NaT
                        plot_data.append({
                            'Product Name': row['Product Name'],
                            'Components': 'Setup Time',  # General label for setup time
                            'Start Time': setup_start,
                            'End Time': row['Start Time'],
                            'Machine Number': '',  # Remove machine number for setup time
                            'Is Setup': True,
                            'Is Late': False,
                            'Delay Days': 0,   # No delay for setup
                            'Delay Hours': 0,  # No delay for setup
                            'Hover Start Time': '',  # Empty string for hover
                            'Hover End Time': ''  # Empty string for hover
                        })

                    # Adjust end time based on the process type of the next row
                    new_end = row['End Time'] - timedelta(minutes=30) if pd.notnull(row['End Time']) else pd.NaT
                    plot_data.append({
                        'Product Name': row['Product Name'],
                        'Components': row['Components'],
                        'Start Time': row['Start Time'],
                        'End Time': new_end,
                        'Machine Number': row['Machine Number'],
                        'Is Setup': False,
                        'Is Late': False,
                        'Delay Days': 0,   # No delay for process components
                        'Delay Hours': 0,  # No delay for process components
                        'Hover Start Time': row['Start Time'].strftime('%d-%b %H:%M') if pd.notnull(row['Start Time']) else '',
                        'Hover End Time': new_end.strftime('%d-%b %H:%M') if pd.notnull(new_end) else ''
                    })

                    last_end_time = row['End Time']  # Update last end time
                    last_machine_number = row['Machine Number']  # Update last machine number

                    # Accumulate delay times if columns exist
                    delay_days = row.get('Delay Days', 0)
                    delay_hours = row.get('Delay Hours', 0)
                    if pd.notnull(delay_days) and (delay_days > 0 or delay_hours > 0):
                        total_delay += timedelta(days=delay_days, hours=delay_hours)
                        delay_count += 1  # Track how many delays

                # Add a single delay block for the total delay at the end of the last component
                if delay_count > 0:
                    delay_days = total_delay.days
                    delay_hours = total_delay.seconds // 3600  # Convert seconds to hours

                    plot_data.append({
                        'Product Name': product_name,
                        'Components': 'Late',  # Change label from Total Delay to Late
                        'Start Time': last_end_time,  # Start Time for late block
                        'End Time': last_end_time + total_delay,  # End Time with delay
                        'Machine Number': last_machine_number,  # Use the last machine number
                        'Is Setup': False,
                        'Is Late': True,  # Indicate that this is a delay block
                        'Delay Days': delay_days,
                        'Delay Hours': delay_hours,
                        'Hover Start Time': 'N/A',  # Display 'N/A' for late block start time
                        'Hover End Time': last_end_time.strftime('%d-%b %H:%M') if pd.notnull(last_end_time) else 'N/A'  # Display end time of the last component for Late block in hover
                    })

            # Convert plot data into DataFrame
            plot_df = pd.DataFrame(plot_data)

            # Drop rows where 'Start Time' is NaT before finding the min start time
            plot_df_clean = plot_df.dropna(subset=['Start Time'])
            min_start_time = plot_df_clean['Start Time'].min().replace(hour=9, minute=0, second=0)

            # Define specific colors for each component and setup time, including delay
            color_discrete_map = {
                "C1": 'skyblue',
                "C2": 'yellow',
                "C3": 'salmon',
                "C4": 'gold',
                "C5": 'orchid',
                "Setup Time": 'black',  # General color for setup times
                "Late": 'red'  # Change color for Late blocks
            }

            # Create the Gantt chart
            fig = px.timeline(
                plot_df,
                x_start='Start Time',
                x_end='End Time',
                y='Product Name',
                color='Components',
                title='Real-Time 2D Gantt Chart',
                labels={'Components': 'Component'},
                color_discrete_map=color_discrete_map,
                custom_data=['Product Name', 'Components', 'Machine Number', 'Hover Start Time', 'Hover End Time', 'Delay Days', 'Delay Hours']
            )

            # Define hovertemplate with conditional display
            fig.update_traces(
                hovertemplate=(
                    "Product Name: %{customdata[0]}<br>"
                    "Component: %{customdata[1]}<br>"
                    "Machine Number: %{customdata[2]}<br>"
                    "Start Time: %{customdata[3]}<br>"  # Display 'N/A' for Late blocks
                    "End Time: %{customdata[4]}<br>"    # Display End Time of the last component for Late blocks
                    "Delay Days: %{customdata[5]}<br>"
                    "Delay Hours: %{customdata[6]}<br>"
                    "<extra></extra>"
                )
            )

            # Update layout to include both date and time on the x-axis
            fig.update_layout(
                xaxis_title="Time",
                yaxis_title="Products",
                xaxis=dict(
                    tickangle=0,
                    tickformat="%d-%b<br>%H:%M",  # Set x-axis format to 'dd-mmm' and 'HH:mm'
                    rangeslider=dict(visible=False),
                    tickmode='linear',
                    tick0=min_start_time,  # Start tick from the minimum Start Time
                    dtick=45000000,  # 5 hours in milliseconds (5 hours * 60 minutes * 60 seconds * 1000 milliseconds)
                    range=[min_start_time, plot_df['End Time'].max()]  # Set the x-axis range to start from min_start_time
                ),
                height=1000,  # Adjust height for better readability
                width=1800,  # Adjust width for better readability
            )

            # Add machine IDs as text inside the rectangles, excluding setup time and late blocks
            for index, row in plot_df.iterrows():
                if pd.notnull(row['Start Time']) and pd.notnull(row['End Time']):
                    start_time = pd.to_datetime(row['Start Time'])
                    end_time = pd.to_datetime(row['End Time'])
                    duration = (end_time - start_time) / 2
                    mid_time = start_time + duration
                    
                    # Only add text for non-setup and non-late components
                    if not row['Is Setup'] and row['Components'] != 'Late':
                        annotation_text = f"{row['Machine Number']}<br>{row['Components']}"
                        fig.add_annotation(
                            x=mid_time.strftime("%Y-%m-%d %H:%M:%S"),
                            y=row['Product Name'],
                            text=annotation_text,
                            showarrow=False,
                            font=dict(color='black', size=9),
                            align='center',
                            xanchor='center',
                            yanchor='middle'
                        )


        elif selected_plot=="Gantt Chart(Unschedule)":
            df1 = fetch_data(full_file_path, "P")
            # Convert the dates to datetime format
            # Convert the dates to datetime format
            df1['Promised Delivery Date'] = pd.to_datetime(df1['Promised Delivery Date'], format='%d-%m-%Y')
            df1['Order Processing Date'] = pd.to_datetime(df1['Order Processing Date'], format='%d-%m-%Y')

            # Create a new column for the list of components
            df1['Component List'] = df1.groupby('Product Name')['Components'].transform(lambda x: ', '.join(x))

            # Define the order of product names
            product_order = df1['Product Name'].unique().tolist()  # or specify your custom order

            # Create the Gantt chart using px.timeline
            fig = px.timeline(
                df1, 
                x_start="Order Processing Date", 
                x_end="Promised Delivery Date", 
                y="Product Name", 
                color="Product Name",  # Color by product name
                hover_data=["Component List", "Order Processing Date", "Promised Delivery Date"],
                title="Gantt Chart: Product Delivery Timeline",
                category_orders={"Product Name": product_order}  # Set the order of product names
            )

            # Add vertical dotted lines and annotate the due date on the x-axis
            for product in df1['Product Name'].unique():
                product_data = df1[df1['Product Name'] == product]
                due_date = product_data['Promised Delivery Date'].max()
                
                # Ensure that due_date is not NaT and a valid datetime
                if pd.notnull(due_date):
                    # Draw a vertical dotted line from the product block to the x-axis (at the due date)
                    fig.add_trace(go.Scatter(
                        x=[due_date, due_date],  # x at the due date
                        y=[df1['Product Name'].min(), df1['Product Name'].max()],  # y from min to max products for full height
                        mode="lines",
                        line=dict(dash='dot', color='black'),
                        hoverinfo='skip',
                        showlegend=False
                    ))
                    
                    # Add an annotation for the due date below the x-axis line
                    fig.add_annotation(
                        x=due_date,
                        y=-0.1,  # Adjust y to place the annotation below the x-axis
                        text=f"{due_date.strftime('%d-%b')}",  # Format the date
                        showarrow=False,
                        xref="x",  # Reference to the x-axis
                        yref="paper",  # This makes the y-coordinate relative to the plot area
                        font=dict(size=10, color="black")
                    )

            # Update the layout to hide x-axis values and ensure all vertical lines appear
            fig.update_layout(
                xaxis_title="Delivery Date",
                yaxis_title="Product Name",
                showlegend=False,  # Hide the legend
                height=600,
                xaxis=dict(
                    showticklabels=False,  # Hide x-axis tick labels
                    showline=False,  # Hide x-axis line
                    showgrid=True,  # Still show the grid
                    zeroline=True  # Optionally show the zero line
                ),
                yaxis=dict(
                    autorange='reversed'  # Reverse the y-axis to show the first product at the top
                )
            )

        elif selected_plot == "Utilization":
            df['Time Diff'] = df['Time Diff'].apply(time_to_timedelta2)
            df['Utilization'] = df['Time Diff'].apply(calculate_utilization)
            #print(df)
            # Convert Machine Number to categorical
            df['Machine Number'] = df['Machine Number'].astype(str)
            # Remove rows where Machine Number is 'OutSrc'
            df = df[df['Machine Number'] != 'OutSrc']

            df['Utilization %'] = (df['Utilization'] / 420) * 100
            fig = px.bar(df, x='Machine Number', y='Utilization %',color='Machine Number',
                        labels={'Utilization %': 'Utilization (%)', 'Machine Number': 'Machine'},
                        title='Utilization Percentage for Each Machine',color_discrete_sequence=['red', 'blue', 'green', 'orange', 'purple'])
            
            # Update y-axis to range from 0 to 100
            fig.update_yaxes(range=[0, 100])
        elif selected_plot == "Time Taken by each Machine":
            df['Time Diff'] = df['Time Diff'].apply(time_to_timedelta2)
            # Remove rows where Machine Number is 'OutSrc'
            df = df[df['Machine Number'] != 'OutSrc']
            total_running_time = df.groupby('Machine Number')['Time Diff'].sum().reset_index()
            y_ticks = pd.to_timedelta(range(0, int(total_running_time['Time Diff'].max().total_seconds()) + 1, 8000), unit='s')
            fig = px.bar(total_running_time, x='Machine Number', y='Time Diff',color='Machine Number',
                        labels={'Time Diff': 'Total Running Time (hh:mm:ss)', 'Machine Number': 'Machine'},
                        title='Total Running Time for Each Machine',color_discrete_sequence=['red', 'blue', 'green', 'orange', 'purple'])
            fig.update_layout(
                yaxis=dict(
                    tickmode='array',
                    tickvals=y_ticks,
                    ticktext=[str(td)[7:] for td in y_ticks]
                )
            )
        elif selected_plot == "Time taken by each product":
            df['Time Diff'] = df['Time Diff'].apply(time_to_timedelta2)
            total_running_time_component = df.groupby(['Product Name', 'Components'])['Time Diff'].sum().reset_index()
            total_running_time_component['Product_Component'] = total_running_time_component['Product Name'] + ' - ' + total_running_time_component['Components']
            max_seconds = int(total_running_time_component['Time Diff'].max().total_seconds())
            y_ticks = pd.to_timedelta(range(0, max_seconds + 1, 7000), unit='s')
            fig = px.bar(total_running_time_component, x='Product_Component', y='Time Diff',
                        labels={'Time Diff': 'Total Running Time (hh:mm:ss)', 'Product_Component': 'Product - Component'},
                        title='Total Running Time for Each Product and Component')
            fig.update_layout(
                yaxis=dict(
                    tickmode='array',
                    tickvals=y_ticks,
                    ticktext=[str(td)[7:] for td in y_ticks]
                ),
                xaxis_tickvals=total_running_time_component['Product_Component'],
                xaxis_ticktext=[f"{p.split(' - ')[0]}\n{p.split(' - ')[1]}" for p in total_running_time_component['Product_Component']]
            )
        elif selected_plot == "Wait Time":
            df['Wait Time'] = df['Wait Time'].apply(time_to_timedelta2)
            total_wait_time_component = df.groupby(['Product Name', 'Components'])['Wait Time'].sum().reset_index()
            total_wait_time_component['Product_Component'] = total_wait_time_component['Product Name'] + ' - ' + total_wait_time_component['Components']
            max_seconds = int(total_wait_time_component['Wait Time'].max().total_seconds())
            y_ticks = pd.to_timedelta(range(0, max_seconds + 1, 5000), unit='s')
            fig = px.bar(total_wait_time_component, x='Product_Component', y='Wait Time',
                        labels={'Wait Time': 'Total Wait Time (hh:mm:ss)', 'Product_Component': 'Product - Component'},
                        title='Total Wait Time for Each Product and Component')
            fig.update_layout(
                yaxis=dict(
                    tickmode='array',
                    tickvals=y_ticks,
                    ticktext=[str(td)[7:] for td in y_ticks]
                ),
                xaxis_tickvals=total_wait_time_component['Product_Component'],
                xaxis_ticktext=[f"{p.split(' - ')[0]}\n{p.split(' - ')[1]}" for p in total_wait_time_component['Product_Component']]
            )
        elif selected_plot == "Idle Time":
            df['Idle Time'] = df['Idle Time'].apply(time_to_timedelta2)
            # Remove rows where Machine Number is 'OutSrc'
            df = df[df['Machine Number'] != 'OutSrc']
            total_ideal_time = df.groupby('Machine Number')['Idle Time'].sum().reset_index()
            y_ticks_1 = pd.to_timedelta(range(0, int(total_ideal_time['Idle Time'].max().total_seconds()) + 1, 15000), unit='s')
            fig = px.bar(total_ideal_time, x='Machine Number', y='Idle Time',
                        labels={'Idle Time': 'Total Idle Time (hh:mm:ss)', 'Machine Number': 'Machine'},
                        title='Total Idle Time for Each Machine',color_discrete_sequence=['red', 'blue', 'green', 'orange', 'purple'])
            fig.update_layout(
                yaxis=dict(
                    tickmode='array',
                    tickvals=y_ticks_1,
                    ticktext=[str(td)[7:] for td in y_ticks_1],
                    tickformat='%H:%M:%S'
                )
            )
        elif selected_plot == "Product Components Status":
            # Clean up the Status column
            df['Status'] = df['Status'].apply(lambda x: str(x).strip() if x is not None else '')

            # Define the colors and custom legends for statuses
            status_colors = {
                "InProgress_Outsource": "orange",
                "InProgress_In House": "yellow",
                "Completed_In House": "green",
                "Completed_Outsource": "blue",
                "Late": "red"  # Use a common color for both Late statuses
            }

            custom_legend_names = {
                "InProgress_Outsource": "Component Out for Outsource",
                "Completed_Outsource": "Component Back From Outsource and Completed",
                "InProgress_In House": "Component InProgress Inhouse",
                "Completed_In House": "Component Completed Inhouse",
                "Late": "Component Late"  # Use the same legend entry for both Late statuses
            }

            fig = go.Figure()

            # Loop through each product and its components to plot them individually
            for product in df['Product Name'].unique():
                product_data = df[df['Product Name'] == product]
                for component in product_data['Components'].unique():
                    component_data = product_data[product_data['Components'] == component]
                    status = component_data['Status'].values[0]
                    process_type = component_data['Process Type'].values[0]
                    machine_number = component_data['Machine Number'].values[0]
                    # Fill missing Delay Days and Delay Hours with 0
                    delay_days = component_data['Delay Days'].fillna(0).values[0]
                    delay_hours = component_data['Delay Hours'].fillna(0).values[0]
                    start_time = component_data['Start Time'].values[0] if 'Start Time' in component_data else None
                    end_time = component_data['End Time'].values[0] if 'End Time' in component_data else None
                    
                    # Convert start and end times to datetime if they are not already
                    start_time = pd.to_datetime(start_time, errors='coerce')
                    end_time = pd.to_datetime(end_time, errors='coerce')
                    
                    # Combine Late statuses into a single key
                    key = f"{status}_{process_type}" if status != "Late" else "Late"
                    color = status_colors.get(key, "grey")

                    # Create hover text content
                    hover_text = (
                        f"Product Name: {product}<br>"
                        f"Component: {component}<br>"
                        f"Machine Number: {machine_number}<br>"
                        f"Delay Days: {delay_days}<br>"
                        f"Delay Hours: {delay_hours}<br>"
                        f"Start Time: {start_time.strftime('%d-%b %H:%M') if pd.notnull(start_time) else 'N/A'}<br>"
                        f"End Time: {end_time.strftime('%d-%b %H:%M') if pd.notnull(end_time) else 'N/A'}"
                    )
                    
                    fig.add_trace(go.Scatter(
                        x=[product],
                        y=[component],
                        mode='markers+text',
                        marker=dict(
                            color=color,
                            size=30,
                            symbol='square'
                        ),
                        text=[machine_number],
                        textposition='middle center',
                        name='',
                        legendgroup=key,
                        showlegend=False,
                        hovertemplate=hover_text  # Adding hover information
                    ))

            # Create legend items manually
            for status_key, color in status_colors.items():
                legend_name = custom_legend_names.get(status_key, status_key.replace("_", " and "))
                fig.add_trace(go.Scatter(
                    x=[None], y=[None],
                    mode='markers',
                    marker=dict(
                        color=color,
                        size=10,
                        symbol='square'
                    ),
                    legendgroup=status_key,
                    showlegend=True,
                    name=legend_name
                ))

            # Update layout
            fig.update_layout(
                title='Status of Each Product Component',
                xaxis_title='Product',
                yaxis_title='Component',
                xaxis=dict(tickmode='array', tickvals=df['Product Name'].unique()),
                yaxis=dict(tickmode='array', tickvals=df['Components'].unique()),
                legend_title_text='Status and Process Type'
            )

        elif selected_plot=="Remaining Time":
            product_times = {}
                
            for product in df['Product Name'].unique():
                product_df = df[df['Product Name'] == product]
                
                total_time = product_df['Run Time (min/1000)'].sum()
                #remaining_time = product_df[product_df['Status'] != 'Completed']['Run Time (min/1000)'].sum()
                remaining_time = product_df[~product_df['Status'].isin(['Completed', 'Late'])]['Run Time (min/1000)'].sum()

                product_times[product] = {
                    'Total Time': total_time,
                    'Remaining Time': remaining_time
                }

            time_df=pd.DataFrame(product_times).T.reset_index().rename(columns={'index': 'Product Name'})
            #print(time_df)


            # Calculate the completed time
            time_df['Completed Time'] = time_df['Total Time'] - time_df['Remaining Time']

            # Create a stacked horizontal bar chart
            fig = go.Figure()

            # Add trace for Completed Time
            fig.add_trace(go.Bar(
                y=time_df['Product Name'],
                x=time_df['Completed Time'],
                orientation='h',
                name='Completed Time',
                marker=dict(color='green'),
                text=time_df['Completed Time'],
                textposition='inside',
            ))

            # Add trace for Remaining Time
            fig.add_trace(go.Bar(
                y=time_df['Product Name'],
                x=time_df['Remaining Time'],
                orientation='h',
                name='Remaining Time',
                marker=dict(color='red'),
                text=time_df['Remaining Time'],
                textposition='inside',
            ))

            # Update layout
            fig.update_layout(
                barmode='stack',
                title='Total and Remaining Time for Each Product',
                xaxis_title='Time (min)',
                yaxis_title='Product Name',
                yaxis=dict(automargin=True),
                legend=dict(
                    orientation='h',
                    yanchor='bottom',
                    y=1.02,
                    xanchor='right',
                    x=1
                ),
                margin=dict(l=0, r=0, t=50, b=0)  # Adjust margin as needed
            )

        return fig

    app.run_server(debug=False)


if __name__ == '__main__':
    launch_dashboard()
    
    
    
