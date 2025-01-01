import pandas as pd
import numpy as np
import time as tm
from datetime import datetime, timedelta, date,time
import os
import plotly.graph_objects as go
import openpyxl
import sys
import queue
from openpyxl import load_workbook
from openpyxl.styles import Font
from collections import defaultdict

# Initialize machine status dictionary
O1 = 0

interrupted = False

def read_all_sheets(file_path):
    try:
        sheets = pd.read_excel(file_path, sheet_name=None)  # Read all sheets into a dictionary
        return sheets
    except Exception as e:
        print(f"Error reading sheets: {e}")
        return {}

def write_excel(df, file_path, sheet_name, runtime=None):
    if runtime is not None:
        runtime_df = pd.DataFrame({'Run_time': [runtime]})
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            runtime_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def fetch_data(file_path, sheet_name='prodet', filter_completed=False):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if filter_completed:
            df = df[df['Status'] == 'Completed']
            if not df.empty:
                df['End Time'] = pd.to_datetime(df['End Time'], format='%Y-%m-%d %H:%M:%S')
                latest_row = df.loc[df['End Time'].idxmax()]
                return latest_row['End Time'], latest_row['Machine Number']
            else:
                return None, None
        else:
            addln_df = pd.read_excel(file_path, sheet_name='Addln')
            if not addln_df.empty:
                df = pd.concat([df, addln_df], ignore_index=True)
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='prodet', index=False)
                    pd.DataFrame().to_excel(writer, sheet_name='Addln', index=False)
            return df
    except Exception as e:
        print(f"Error fetching data: {e}")
        return pd.DataFrame()

# Define working hours and working days
WORK_START = 9  # 9 AM
WORK_END = 17 + 1/60  # 5 PM
WEEKENDS = [5, 6]  # Saturday and Sunday
machine_schedule = defaultdict(list)
machine_last_end = {}

# Function to adjust end time for working hours and working days
def adjust_to_working_hours_and_days(start_time, run_time_minutes):
    DAILY_WORK_MINUTES = (WORK_END - WORK_START) * 60  # Convert hours to minutes
    current_time = start_time
    remaining_minutes = run_time_minutes

    while remaining_minutes > 0:
        if current_time.hour < WORK_START or current_time.weekday() in WEEKENDS:
            current_time = next_working_day(current_time.replace(hour=WORK_START, minute=0))
        available_minutes_today = max(0, (WORK_END - current_time.hour) * 60 - current_time.minute - 1)
        if remaining_minutes <= available_minutes_today:
            current_time += timedelta(minutes=remaining_minutes)
            remaining_minutes = 0
        else:
            remaining_minutes -= available_minutes_today
            current_time = current_time.replace(hour=WORK_START, minute=0) + timedelta(days=1)
            current_time = next_working_day(current_time)
    return current_time

# Function to find gaps in the machine schedule
def find_gaps(machine_schedule):
    gaps = {}
    for machine, tasks in machine_schedule.items():
        tasks = sorted(tasks, key=lambda x: x[0])  # Sort tasks by start time
        gaps[machine] = []
        for i in range(len(tasks) - 1):
            current_end = tasks[i][1]
            next_start = tasks[i + 1][0]
            if next_start > current_end:
                gaps[machine].append((current_end, next_start))  # Record gap
    return gaps

# Main scheduling function with gap optimization
def schedule_production_with_days(data, machine_schedule, machine_last_end):
    for idx, row in data.iterrows():
        component = row['Components']
        product = row['Product Name']
        machine = row['Machine Number']

        # Determine start time
        if "C1" in component and machine == "OutSrc":
            start_time = next_working_day(row['Order Processing Date'].replace(hour=WORK_START, minute=0))
        else:
            prev_component = f"C{int(component[1:]) - 1}"
            same_product_prev = data[(data['Product Name'] == product) & (data['Components'] == prev_component)]
            if not same_product_prev.empty:
                start_time = same_product_prev.iloc[0]['End Time']
            else:
                gaps = find_gaps(machine_schedule)
                for gap_start, gap_end in gaps.get(machine, []):
                    run_time_minutes = row['Run Time (min/1000)'] * row['Quantity Required'] / 1000
                    potential_end_time = adjust_to_working_hours_and_days(gap_start, run_time_minutes)
                    if potential_end_time <= gap_end:  # Check if the task fits in the gap
                        start_time = gap_start
                        break
                else:
                    start_time = max(machine_last_end[machine], next_working_day(row['Order Processing Date'].replace(hour=WORK_START, minute=0)))

            # Adjust start time to working hours
            if start_time.hour < WORK_START:
                start_time = start_time.replace(hour=WORK_START, minute=0)
            elif start_time.hour >= WORK_END or start_time.weekday() in WEEKENDS:
                start_time = next_working_day(start_time.replace(hour=WORK_START, minute=0))

        # Calculate runtime and end time
        run_time_minutes = row['Run Time (min/1000)'] * row['Quantity Required'] / 1000
        end_time = adjust_to_working_hours_and_days(start_time, run_time_minutes)

        # Update machine schedule
        if machine != "OutSrc":
            machine_schedule[machine].append((start_time, end_time, idx))
            machine_schedule[machine] = sorted(machine_schedule[machine], key=lambda x: x[0])
            machine_last_end[machine] = max(machine_last_end[machine], end_time)

        # Assign start and end times to the DataFrame
        data.loc[idx, 'Start Time'] = start_time
        data.loc[idx, 'End Time'] = end_time

    return data

# Function to adjust End Time if it equals {date} 09:00:00
def adjust_end_time(data):
    for idx, row in data.iterrows():
        if row['End Time'].hour == 9 and row['End Time'].minute == 0 and row['End Time'].second == 0:
            data.at[idx, 'End Time'] = row['End Time'] - timedelta(days=1) + timedelta(hours=17)
    return data

# Function to find the next valid working day
def next_working_day(current_date):
    while current_date.weekday() in WEEKENDS:  # Check if the day is a weekend
        current_date += timedelta(days=1)  # Move to the next day
    return current_date

def setup_time_check(row, similarity_lookup):
    product = row["Product Name"]
    machine = row["Machine Number"]
    if machine == "OutSrc":
        return 1
    if machine in similarity_lookup and product in similarity_lookup[machine]:
        return similarity_lookup[machine][product]
    return 0

def apply_formatting(workbook, sheet_name, df, condition_col, condition_value, font_style):
    ws = workbook[sheet_name]
    for row_idx, row in df.iterrows():
        if row[condition_col] == condition_value:
            excel_row = row_idx + 2
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=excel_row, column=col).font = font_style

def process_initialization(file_path):
    print("Initialization process started.")
    sheets = read_all_sheets(file_path)
    components_df = sheets.get('prodet', pd.DataFrame())
    # Convert columns to appropriate types
    components_df['Promised Delivery Date'] = pd.to_datetime(components_df['Promised Delivery Date'])
    components_df['Ready Time'] = pd.to_datetime(components_df['Ready Time'])
    components_df['Start Time'] = pd.NaT  # Initialize as empty datetime
    components_df['End Time'] = pd.NaT  # Initialize as empty datetime

    # Sort the data by Promised Delivery Date, Product Name, and Component order
    components_df = components_df.sort_values(by=['Promised Delivery Date',
                        'Product Name',
                        'Components']).reset_index(drop=True)

    outsource_df = sheets.get('Outsource Time', pd.DataFrame())
    machines_df = sheets.get('Machines', pd.DataFrame())
    similarity_df = sheets.get('Similarity', pd.DataFrame())

    similarity_lookup = similarity_df.set_index('Machine').T.to_dict()
    components_df['SetupTimeCheck'] = components_df.apply(lambda row: setup_time_check(row, similarity_lookup), axis=1)

    # Initialize machine_last_end
    machine_last_end = {
        machine: next_working_day(
            components_df['Order Processing Date'].min().replace(hour=WORK_START, minute=0)
        )
        for machine in components_df['Machine Number'].unique()
    }

    # Call schedule_production_with_days with shared state
    components_df = schedule_production_with_days(components_df, machine_schedule, machine_last_end)

    write_excel(components_df, file_path, 'prodet')
    return components_df, outsource_df, machines_df, similarity_df

def process_start(file_path):
    print("Start process initiated.")
    outsource_df = pd.read_excel(file_path, sheet_name='Outsource Time')
    machines_df = pd.read_excel(file_path, sheet_name='Machines')
    similarity_df = pd.read_excel(file_path, sheet_name='Similarity')
    
    components_df = fetch_data(file_path)
    # Convert columns to appropriate types
    components_df['Promised Delivery Date'] = pd.to_datetime(components_df['Promised Delivery Date'])
    components_df['Ready Time'] = pd.to_datetime(components_df['Ready Time'])
    components_df['Start Time'] = pd.NaT  # Initialize as empty datetime
    components_df['End Time'] = pd.NaT  # Initialize as empty datetime

    # Sort the data by Promised Delivery Date, Product Name, and Component order
    components_df = components_df.sort_values(by=['Promised Delivery Date',
                        'Product Name',
                        'Components']).reset_index(drop=True)

    # Convert DataFrame to dictionary for similarity lookup
    similarity_lookup = similarity_df.set_index('Machine').T.to_dict()
    components_df['SetupTimeCheck'] = components_df.apply(lambda row: setup_time_check(row, similarity_lookup), axis=1)

    # Call schedule_production_with_days directly for scheduling
    components_df = schedule_production_with_days(components_df)

    return components_df, outsource_df, machines_df, similarity_df

def calculate_and_format_data(components_df):
    # Convert 'Start Time' and 'End Time' columns to datetime
    components_df['Start Time'] = pd.to_datetime(components_df['Start Time'])
    components_df['End Time'] = pd.to_datetime(components_df['End Time'])

    # Calculate the duration (difference) between start and end times
    components_df['Time Diff_days'] = components_df['End Time'] - components_df['Start Time']

    # Fixed total time (start time in this case)
    total_time = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)

    # Calculate Idle Time
    components_df['Idle Time'] = total_time - components_df['Time Diff_days']

    # Format Idle Time to %H:%M:%S
    components_df['Idle Time'] = components_df['Idle Time'].dt.strftime('%H:%M')

    # Step 1: Convert the time difference from timedelta to total seconds
    components_df['Time Diff'] = components_df['Time Diff_days'].dt.total_seconds()

    # Step 2: Convert total seconds to hours and minutes format '%H:%M'
    components_df['Time Diff'] = components_df['Time Diff'].apply(
        lambda x: pd.to_datetime(x, unit='s').strftime('%H:%M:%S')
    )

    components_df['Promised Delivery Date'] = components_df['Promised Delivery Date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    components_df['Order Processing Date'] = components_df['Order Processing Date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    components_df['Start Time'] = components_df['Start Time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    components_df['End Time'] = components_df['End Time'].dt.strftime('%Y-%m-%d %H:%M:%S')

    return components_df

def apply_highlight_formatting(file_path, components_df):
    """
    Apply red font to rows in the 'prodet' sheet where 'Status' is 'Late'.
    """
    wb = load_workbook(file_path)
    ws = wb['prodet']  # Assuming the sheet name is 'prodet'
    red_font = Font(color="FF0000", bold=True)

    # Iterate through the rows in components_df
    for index, row in components_df.iterrows():
        if row['Status'] == 'Late':  # Check the condition
            excel_row = index + 2  # Adjust for header row
            for col_idx, value in enumerate(row, start=1):  # Loop through all columns
                ws.cell(row=excel_row, column=col_idx).font = red_font

    # Save the workbook after applying formatting
    wb.save(file_path)


def main():
    file_path = r'Product Details_v2.xlsx'
    print(file_path)
    input_data = sys.stdin.read().strip()
    print(input_data)

    if input_data == "Initial":
        components_df, outsource_df, machines_df, similarity_df = process_initialization(file_path)
    elif input_data == "Start":
        components_df, outsource_df, machines_df, similarity_df = process_start(file_path)
    else:
        print(f"Unknown command: {input_data}")
        return


    # Process and format data
    # components_df = calculate_and_format_data(components_df)
    components_df = schedule_production_with_days(components_df)
    components_df = adjust_end_time(components_df)

    # Save the results to the Excel file
    write_excel(components_df, file_path, 'prodet')
    write_excel(outsource_df, file_path, 'Outsource Time')
    write_excel(machines_df, file_path, 'Machines')
    write_excel(similarity_df, file_path, 'Similarity')

    # Apply formatting for "Late" status
    apply_highlight_formatting(file_path, components_df)


if __name__ == "__main__":
    main()
