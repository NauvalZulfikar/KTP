import streamlit as st
from scheduler import dfm
import pandas as pd
from openpyxl import load_workbook

def save_to_excel(df, sheet_name):
    try:
        # Load the existing workbook
        book = load_workbook(file_path)

        # Open the workbook with ExcelWriter
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            writer.book = book  # Associate the writer with the loaded workbook
            writer.sheets = {ws.title: ws for ws in book.worksheets}  # Load existing sheets
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write to the sheet
    except Exception as e:
        st.error(f"Error saving data to Excel: {e}")

file_path = "Product Details_v1.xlsx"

# Load the dataframe
dfn = dfm.drop(columns=['wait_time', 'legend', 'Status']).copy()

def save_to_excel(df, sheet_name):
    """Save the updated dataframe back to the Excel file."""
    try:
        book = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        st.error(f"Error saving data to Excel: {e}")

def modify():
    # Add Tabs Below
    tabs = st.tabs([
        "In House", 
        "Out Source", 
        "Time Converter"
        ])

    int_col = ['UniqueID', 'Sr. No', 'Quantity Required', 'Run Time (min/1000)', 'Cycle Time (seconds)', 'Setup time (seconds)']
    str_col = ['Product Name', 'Components', 'Operation', 'Process Type', 'Machine Number']
    date_col = ['Order Processing Date', 'Promised Delivery Date']

    with tabs[0]:  # In House
        df_in = dfn[dfn['Process Type'] == 'In House']
        in_products = df_in['Product Name'].unique()
        in_selected_product = st.selectbox(
            'Select product name:',
            in_products,
            key="in_product"
        )

        in_components = df_in[df_in['Product Name'] == in_selected_product]['Components'].unique()
        in_selected_components = st.selectbox(
            'Select components:',
            in_components,
            key="in_component"
        )

        in_field = df_in.columns
        in_selected_fields = st.selectbox(
            'Select fields:',
            in_field,
            key="in_field"
        )

        if in_selected_fields in int_col:
            in_edit_input = st.number_input(
                'Enter new value:',
                key="in_edit_input"
            )
        elif in_selected_fields in str_col:
            in_edit_input = st.text_input(
                'Enter new value:',
                key="in_edit_text"
            )
        else:
            in_edit_input = st.date_input(
                'Enter new value:',
                key="in_edit_date"
            )
            
        if st.button('Confirm', key="in_confirm"):
            df_in.loc[
                (df_in['Product Name'] == in_selected_product) &
                (df_in['Components'] == in_selected_components),
                in_selected_fields
            ] = in_edit_input

            # Save changes back to Excel
            save_to_excel(dfn, sheet_name="prodet")
        
        st.dataframe(df_in[
            (df_in['Product Name'] == in_selected_product) &
            (df_in['Components'] == in_selected_components)
        ])

    with tabs[1]:  # Outsource
        df_out = dfn[dfn['Process Type'] == 'Outsource']
        out_products = df_out['Product Name'].unique()
        out_selected_product = st.selectbox(
            'Select product name:',
            out_products,
            key="out_product"
        )

        out_components = df_out[df_out['Product Name'] == out_selected_product]['Components'].unique()
        out_selected_components = st.selectbox(
            'Select components:',
            out_components,
            key="out_component"
        )

        out_field = df_out.columns
        out_selected_fields = st.selectbox(
            'Select fields:',
            out_field,
            key="out_field"
        )
        
        if out_selected_fields in int_col:
            out_edit_input = st.number_input(
                'Enter new value:',
                key="out_edit_input"
            )
        elif out_selected_fields in str_col:
            out_edit_input = st.text_input(
                'Enter new value:',
                key="out_edit_text"
            )
        else:
            out_edit_input = st.date_input(
                'Enter new value:',
                key="out_edit_date"
            )
            
        if st.button('Confirm', key="out_confirm"):
            df_out.loc[
                (df_out['Product Name'] == out_selected_product) &
                (df_out['Components'] == out_selected_components),
                out_selected_fields
            ] = out_edit_input

            # Save changes back to Excel
            save_to_excel(dfn, sheet_name="prodet")
        
        st.dataframe(df_out[
            (df_out['Product Name'] == out_selected_product) &
            (df_out['Components'] == out_selected_components)
        ])
        
    with tabs[2]:  # Time Converter
        # Radio button for conversion options
        conversion_type = st.radio(
            "Choose a conversion type:",
            ("Days to Minutes", "Hours to Minutes", "Minutes to Days"),
            key="conversion_type"
        )
        
        # Input field for the user to provide a value
        input_value = st.number_input(
            "Enter the value to convert:", 
            min_value=0.0, 
            step=1.0,
            format="%.2f",
            key="conversion_input"
        )
        
        # Perform conversion based on the selected type
        if conversion_type == "Days to Minutes":
            result = input_value * 24 * 60
            st.write(f"{input_value} days is equivalent to {result} minutes.")
        
        elif conversion_type == "Hours to Minutes":
            result = input_value * 60
            st.write(f"{input_value} hours is equivalent to {result} minutes.")
        
        elif conversion_type == "Minutes to Days":
            result = input_value / (24 * 60)
            st.write(f"{input_value} minutes is equivalent to {result:.6f} days.")
